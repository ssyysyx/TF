#!/usr/bin/env python
# coding: utf-8

# ## 打开文件

# In[1]:


print("------------------------检测万得登录情况：------------------------")


# In[2]:


from WindPy import *
import xlwings as xw
import numpy as np
import os
import pandas as pd
from openpyxl import load_workbook
import warnings
import datetime
warnings.filterwarnings('ignore')
w.start()
print("------------------------万得已登录------------------------")

path = os.getcwd()
data_file = os.listdir(path + '/data')
result_file = os.listdir(path + '/result' )

print("***************************************************************")
print("Input File List:")
print(path+'/data/组合证券')
print(path+'/data/投资组合')
print(path+'/data/债券持仓')

print("****************************************************************")
print("Find and Open:")

for f in data_file:
    if "组合证券" in f:
        print(path + '/data/' + f)
        df_jiaoyisuo_records = pd.read_excel(path + '/data/' + f,sheet_name=0,dtype={'证券代码': str,'持仓':float})
    if "投资组合" in f:
        print(path + '/data/' + f)
        df_yinhangjian_records = pd.read_excel(path + '/data/' + f,sheet_name=0,dtype={'面额': float})
    if "债券持仓" in f:
        filename_ori = path + '/data/' + f
        print(path + '/data/' + f)
        df_result_records = pd.read_excel(path + '/data/' + f,sheet_name='债券类持有列表',dtype={'代码': str,'面额': float})
        wb = xw.Book(path + '/data/' + f)

print("-----------------------忽略警告【warning】----------------------")

td = datetime.date.today()
tod = td
ytd = w.tdaysoffset(-1, td.strftime("%Y-%m-%d"), "").Data[0][0]
td = str(td)
riqi = td[5:7] + td[8:10]
file_name = path + '/result/债券持仓情况集中度'+riqi+'.xlsx'


# In[4]:


last_row = int(wb.sheets['总持仓变化'].range('A' + str(wb.sheets[1].cells.last_cell.row)).end('up').row)
wb.sheets['总持仓变化'].range('A'+str(last_row+1)).value = tod


# ## 数据清洗

# In[5]:


# 只留下资产类别列包含债券资产的行
df_jiaoyisuo_records = df_jiaoyisuo_records[df_jiaoyisuo_records['资产类别']=='债券资产']
# 简化列
df_jiaoyisuo_records = df_jiaoyisuo_records[['证券代码','持仓','证券名称','交易市场']]
df_yinhangjian_records = df_yinhangjian_records[['Unnamed: 2','面额']]
df_yinhangjian_records['证券名称'] = pd.Series(df_yinhangjian_records['Unnamed: 2'],index = df_yinhangjian_records.index)
# 将带有空值的行全部删除
df_yinhangjian_records = df_yinhangjian_records.dropna()
df_jiaoyisuo_records = df_jiaoyisuo_records.dropna()
# 删除Unnamed: 2这列包含某些关键字的行
df_yinhangjian_records = df_yinhangjian_records[~df_yinhangjian_records['Unnamed: 2'].str.contains('质押式回购')]
df_yinhangjian_records = df_yinhangjian_records[~df_yinhangjian_records['Unnamed: 2'].str.contains('拆入')]
# 提取括号内的内容
df_yinhangjian_records['Unnamed: 2'] = df_yinhangjian_records["Unnamed: 2"].str.replace(r'.*\(|\).*', '')
# 提取括号外的内容
df_yinhangjian_records['证券名称'] = df_yinhangjian_records['证券名称'].str.replace(r"\(.*\)","")


# ## 更新持仓信息

# In[6]:


# arrs保存代码列的所有值
# print(type(df_result_records.loc[2,'代码']))
col = df_result_records.iloc[:,1]
arrs = col.values


# In[7]:


SZ = '深交所A'
SH = '上交所A'
df_result_records['面额'] = 0.00

# 更新交易所持仓信息
for tup in df_jiaoyisuo_records.itertuples():
    mytup1 = tup[1::]
    if mytup1[0] in arrs:
        val_init = df_result_records.loc[df_result_records['代码'] == mytup1[0],'面额']
        val_init = val_init.values
        if val_init == 0.00:
            df_result_records.loc[df_result_records['代码'] == mytup1[0],'面额'] = mytup1[1]/100.0
        else:
            hang = df_result_records.loc[df_result_records['代码'] == mytup1[0]]
            df_result_records.loc[df_result_records.shape[0]+1] = [hang['场所'].values[0],hang['代码'].values[0],hang['债券标的'].values[0],hang['行业'].values[0],mytup1[1]/100.0]

    else:
        if mytup1[3] == SZ:
            sql_code = str(mytup1[0]+'.SZ')
        else:sql_code = str(mytup1[0]+'.SH')
        trans_industry = w.wsd(sql_code, "industry_csrc12_n", "ED0D", ytd.strftime("%Y-%m-%d"), "industryType=1").Data[0][0]
        df_result_records.loc[df_result_records.shape[0]+1] = [mytup1[3],mytup1[0],mytup1[2],trans_industry,mytup1[1]/100.0 ]

# 更新银行间持仓信息
for tup in df_yinhangjian_records.itertuples():
    mytup1 = tup[1::]
    if mytup1[0] in arrs:
        val_init = df_result_records.loc[df_result_records['代码'] == mytup1[0],'面额']
        val_init = val_init.values
        if val_init == 0.00:
            df_result_records.loc[df_result_records['代码'] == mytup1[0],'面额'] = mytup1[1]/10000.0
        else:
            hang = df_result_records.loc[df_result_records['代码'] == mytup1[0]]
            df_result_records.loc[df_result_records.shape[0]+1] = [hang['场所'].values[0],hang['代码'].values[0],hang['债券标的'].values[0],hang['行业'].values[0],mytup1[1]/10000.0]
    else:
        sql_code = mytup1[0] + '.IB'
        trans_industry = w.wsd(sql_code, "industry_csrc12_n", "ED0D", ytd.strftime("%Y-%m-%d"), "industryType=1").Data[0][0]
        df_result_records.loc[df_result_records.shape[0]+1] = ['银行间',mytup1[0],mytup1[2],trans_industry,mytup1[1]/10000.0 ]


# In[8]:


# 更新占比表
df_cate = pd.pivot_table(df_result_records,index = ["行业"],values = ["面额"],aggfunc = np.sum)
df_cate['占比'] = df_cate['面额']/df_cate['面额'].sum()

df_cate = df_cate.sort_values(by = '占比',ascending = False)

def turn_percentage(x):
    return '%.2f%%' % (x* 100)
df_cate['占比'] = df_cate['占比'].apply(turn_percentage)

# df_cate['占比'] = df_cate['占比'].apply(lambda x:'%.2f%%' % (x* 100) ,axis=1)

df_cate.loc["持仓总和"] = df_cate.apply(lambda x:x.sum())
df_cate.loc["持仓总和","占比"] = '100%'

df_cate = df_cate.round({'面额': 1})


# ## 保存至result

# In[9]:


wb.sheets['总持仓变化'].range('B'+str(last_row+1)).value = df_cate.loc['持仓总和','面额']
wb.sheets['债券类持仓'].clear()
wb.sheets['占比'].clear()
wb.sheets['房地产类持仓'].clear()


# In[10]:


df_cate['面额'] = df_cate['面额'].astype(str)
df_state = df_result_records[df_result_records['行业']=='房地产业']

wb.sheets['债券类持仓'].range('A1').expand('table').value = df_result_records
wb.sheets['占比'].range('A1').expand('table').value = df_cate
wb.sheets['房地产类持仓'].range('A1').expand('table').value = df_state

wb.sheets['债券类持仓']['A:A'].delete()
wb.sheets['房地产类持仓']['A:A'].delete()

wb.sheets['占比'].range("A1:C1").api.Font.Bold = True


# In[11]:


wb.save(file_name)
wb.close()

print("****************************************************************")
print("Output File:")
print(file_name)
print("****************************************************************")


# In[12]:


# writer = pd.ExcelWriter(file_name,engine = 'openpyxl')
# book = load_workbook(file_name)
# writer.book = book

# df_result_records.to_excel(writer,sheet_name = '债券类持仓',index = False)
# df_cate.to_excel(writer,sheet_name = '占比')
# df_state.to_excel(writer,sheet_name = '房地产类持仓',index = False)
# writer.save()


# In[13]:


# # 调整格式
# from openpyxl.styles import Font
# from openpyxl.styles import Alignment
# workbook = load_workbook(filename = file_name)
# sheet = workbook['占比']
# sheet.column_dimensions['A'].width = 40
# sheet.column_dimensions['B'].width = 15
# sheet.column_dimensions['C'].width = 15
# font = Font(color="FF0000")
# cellss = sheet['C']
# for item in cellss:
#     if item.value!='占比' and item.value!='100%':
#         val = float(item.value[0:4])
#         if val>10:
#             item.font = font
# cells = sheet['C']
# for cell in cells:
#     cell.alignment = Alignment(horizontal = 'right')
# cells = sheet['B']
# for cell in cells:
#     cell.alignment = Alignment(horizontal = 'right')
# cells = sheet['A']
# for cell in cells:
#     cell.alignment = Alignment(horizontal = 'right')
    
# sheet1 = workbook['债券类持仓']
# sheet1.column_dimensions['A'].width = 10
# sheet1.column_dimensions['B'].width = 10
# sheet1.column_dimensions['C'].width = 23
# sheet1.column_dimensions['D'].width = 32
# sheet1.column_dimensions['E'].width = 10
# cells = sheet1['B']
# for cell in cells:
#     cell.alignment = Alignment(horizontal = 'left')
# sheet1.auto_filter.ref = sheet1.dimensions

# sheet2 = workbook['房地产类持仓']
# sheet2.column_dimensions['B'].width = 14
# sheet2.column_dimensions['C'].width = 24
# sheet2.auto_filter.ref = 'A1'
# workbook.save(filename = file_name)

# print("****************************************************************")
# print("Output File:")
# print(file_name)
# print("****************************************************************")


# In[ ]:




