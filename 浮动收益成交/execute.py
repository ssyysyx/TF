#!/usr/bin/env python
# coding: utf-8

# In[1]:


print("------------------------检测万得登录情况：------------------------")
from WindPy import *
w.start()
print("------------------------万得已登录------------------------")
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import datetime


# In[2]:


# 打开浮动收益表文件
path = os.getcwd()
data_file = os.listdir(path+'/data')

print("***************************************************************")
print("Input File List:")
print(path+'/data/浮动收益表')
print(path+'/data/持仓地产债境内成交跟踪')

flag_file_1 = False
flag_file_2 = False

print("***************************************************************")
print("Find and Open:")
for f in data_file:
    if "浮动收益" in f:
        print(path + '/data/' + f)
        file_name = path + '/data/' + f
        workbook = load_workbook(filename = file_name)
        flag_file_1 = True
        
    if "持仓地产债境内成交跟踪" in f:
        print(path + '/data/' + f)
        file_name1 = path + '/data/' + f
        workbook1 = load_workbook(filename = file_name1)
        flag_file_2 = True

if flag_file_1 and flag_file_2:
    print("成功打开文件，正在计算")
elif flag_file_1 == False and flag_file_2:
    print("未找到<浮动收益表>")
elif flag_file_2 == False and flag_file_1:
    print("未找到<持仓地产债境内成交跟踪>")
else:
    print("未找到<浮动收益表>及<持仓地产债境内成交跟踪>")

print("-----------------------------------------------------------------")

for name in workbook.sheetnames:
    if "市值法" in name:
        myname = name
        
sheet = workbook[myname]        
sheet1 = workbook1['境内成交']
#插入四列
sheet.insert_cols(idx = 12, amount = 4)


# In[3]:


# 修改L1:O2的颜色
cell = sheet['J1']
pinkrgb = cell.fill.fgColor
cell = sheet['J2']
yellowrgb = cell.fill.fgColor
pink_list = ['L1','M1','N1','O1']
yellow_list = ['L2','M2','N2','O2']
for item in pink_list:
    sheet[item].fill = PatternFill(fill_type = 'solid', fgColor = pinkrgb)
for item in yellow_list:
    sheet[item].fill = PatternFill(fill_type = 'solid', fgColor = yellowrgb)


# In[4]:


# 修改列名
sheet['L2'] = "当日成交价"
sheet['M2'] = "成交价-净价"
sheet['N2'] = "市场成交笔数"
sheet['O2'] = "收益情况"
# 设置列名居中
alignment = Alignment(horizontal='center')
sheet['L2'].alignment = alignment
sheet['M2'].alignment = alignment
sheet['N2'].alignment = alignment
sheet['O2'].alignment = alignment
# 设置字号
font = Font(name="宋体",size = 10)
sheet['L2'].font = font
sheet['M2'].font = font
sheet['N2'].font = font
sheet['O2'].font = font


# In[5]:


# 从万德拉取当日成交价
td = datetime.date.today()
# oneday = datetime.timedelta(days = 1)
# yd = td - oneday
# str_td = td.strftime("%Y/%m/%d")
# str_d = yd.strftime("%Y/%m/%d")
ytd = w.tdaysoffset(-1, td.strftime("%Y,%m,%d"), "").Data[0][0]
str_d = ytd.strftime("%Y/%m/%d")
print("成交价日: "+ str_d)


# In[6]:


# L列
alignment = Alignment(horizontal = 'right')
font = Font(name="宋体",size = 9)
for cell in sheet['L']:
    if sheet['C'+str(cell.row)].value=="基金代码" or sheet['B'+str(cell.row)].value=="产品名称": 
        num_row = cell.row
        break
    if cell.row>=3:
        if sheet['C'+str(cell.row)].value:
            cell.font = font
            cell.alignment = alignment
            cell.number_format = '###,###,##0.0000'
            cell.value = '=f_dq_close(C'+str(cell.row)+',"'+str_d+'",1)'
#     print(col.value)
# sheet['L3'] = '=f_dq_close(C3,"'+str_d+'",1)'


# In[7]:


# M列
for cell in sheet['M']:
    if cell.row>=3 and cell.row<num_row:
        if sheet['L'+str(cell.row)].value:
            cell.font = font
            cell.alignment = alignment
            cell.number_format = '#,##0.00_);[Red]\(#,##0.00\)'
            cell.value = '=L'+str(cell.row)+'-I'+str(cell.row)


# In[8]:


# N列
alignment = Alignment(horizontal='center')
for cell in sheet['N']:
    if cell.row>=3 and cell.row<num_row:
        if sheet['M'+str(cell.row)].value:
            cell.font = font
            cell.alignment = alignment
            cell.number_format = '0'
            cell.value = '=s_dq_dealnum(C'+str(cell.row)+',"'+str_d+'")'


# In[9]:


# O列
for cell in sheet['O']:
    if cell.row>=3 and cell.row<num_row:
        if sheet['L'+str(cell.row)].value:
            before_value = sheet['K'+str(cell.row)].value
            if before_value:
                before_value = before_value.replace('I','L')
            cell.value = before_value
            cell.font = font
            cell.number_format = '#,##0.00_);[Red]\(#,##0.00\)'
            cell.alignment = alignment
    elif cell.row>=num_row:
        cell.value = sheet['K'+str(cell.row)].value
        cell.number_format = sheet['K'+str(cell.row)].number_format
        cell.font = Font(name="宋体",size = 10)


# In[10]:


# P列
for cell in sheet['P']:
    if cell.row>=3 and cell.row<num_row:
        if cell.value:
            cell.value = str(cell.value).replace('P','T')


# In[11]:


# S列
for cell in sheet['S']:
    if cell.value:
        cell.value = str(cell.value).replace('P','T')
        cell.value = str(cell.value).replace('N','R')
        cell.value = str(cell.value).replace('M','Q')
        
# S1
sheet['S1'].value = sheet['S1'].value.replace("SUQTRODUCT","SUMPRODUCT")
sheet['S1'].value = sheet['S1'].value.replace("SUQ","SUM")


# In[12]:


# O1 G1 I1 K1 Q1
mrow = sheet.max_row
sheet['O1'].value = sheet['E1'].value.replace('K','O')
# sheet['O1'].value = '=SUM(O3:O'+str(mrow)+')'
sheet['O1'].font = Font(name="宋体",color='FFFF0000',bold = True)
sheet['O1'].number_format = '0.0000' 
sheet['G1'].value = sheet['G1'].value.replace('L','P')
sheet['I1'].value = sheet['I1'].value.replace('O','S')
sheet['K1'].value = sheet['K1'].value.replace('R','V')
sheet['K1'].value = sheet['K1'].value.replace('N','R')
sheet['K1'].value = sheet['K1'].value.replace('PVO','PRO')
sheet['Q1'].value = sheet['Q1'].value.replace('Q','U')


# In[13]:


# 加汇总的三行
if sheet['E'+str(mrow)].value:
    index_row = mrow+2
else:
    index_row = mrow
    while(sheet['E'+str(index_row)].value is None):
        index_row-=1
    index_row += 4

font = Font(name="宋体",size = 10 )
sheet['E'+str(index_row)].value = "浮动净价收益"
sheet['E'+str(index_row+1)].value = "成交价的收益"
sheet['E'+str(index_row+2)].value = "隐形风险"
sheet['F'+str(index_row)].value = sheet['E1'].value
sheet['F'+str(index_row+1)].value = sheet['O1'].value
sheet['F'+str(index_row+2)].value = "=F"+str(index_row+1)+"-F"+str(index_row)

sheet['E'+str(index_row)].font = font
sheet['E'+str(index_row+1)].font = font
sheet['E'+str(index_row+2)].font = font
sheet['F'+str(index_row)].font = font
sheet['F'+str(index_row+1)].font = font
sheet['F'+str(index_row+2)].font = font

sheet['F'+str(index_row)].number_format = '0.00'
sheet['F'+str(index_row+1)].number_format = '0.00'
sheet['F'+str(index_row+2)].number_format = '0.00'

patternfill = PatternFill(fill_type = 'solid', fgColor = "e6e6fa")
sheet['E'+str(index_row+2)].fill = patternfill
sheet['F'+str(index_row+2)].fill = patternfill


# In[14]:


def func1(item):
    item = item.replace('：','，')
    item = item.replace(':','，')
    item = item.replace(',','，')
    small_list = [i.strip() for i in item.split('，')]
    small_list = [i for i in small_list if i != '']
    
    dest_str = None
    
    for item in small_list:
        if "成交价" in item:
            dest_str = item
            break
            
    if dest_str == None:
        for item in small_list:
            if "成交在" in item:
                dest_str = item
                break
    
#     print(dest_str)
    if re.findall('[0-9]*\.?[0-9]+-[0-9]*\.?[0-9]+元',dest_str):
        t = re.findall('[0-9]*\.?[0-9]+-[0-9]*\.?[0-9]+元',dest_str)[0]
        numbs = re.findall('[0-9]*\.?[0-9]+',t)
        numbs = [float(i) for i in numbs ]
        if len(numbs)==2:
            dest_str = str((numbs[0]+numbs[1])/2.0)

    elif re.findall('[0-9]*\.?[0-9]+-[0-9]*\.?[0-9]+之间',dest_str):
        t = re.findall('[0-9]*\.?[0-9]+-[0-9]*\.?[0-9]+之间',dest_str)[0]
        numbs = re.findall('[0-9]*\.?[0-9]+',t)
        numbs = [float(i) for i in numbs ]
        if len(numbs)==2:
            dest_str = str((numbs[0]+numbs[1])/2.0) 
            
    elif re.findall('[0-9]*\.?[0-9]+元-[0-9]*\.?[0-9]+元',dest_str):
        t = re.findall('[0-9]*\.?[0-9]+元-[0-9]*\.?[0-9]*\.?[0-9]+元',dest_str)[0]
        numbs = re.findall('[0-9]*\.?[0-9]+',t)
        numbs = [float(i) for i in numbs ]
        if len(numbs)==2:
            dest_str = str((numbs[0]+numbs[1])/2.0)

    elif re.findall('[0-9]*\.?[0-9]+元和[0-9]*\.?[0-9]+元',dest_str):
        t = re.findall('[0-9]*\.?[0-9]+元和[0-9]*\.?[0-9]+元',dest_str)[0]
        numbs = re.findall('[0-9]*\.?[0-9]+',t)
        numbs = [float(i) for i in numbs ]
        if len(numbs)==2:
            dest_str = str((numbs[0]+numbs[1])/2.0)
            
    elif re.findall('[0-9]*\.?[0-9]+元',dest_str):
        t = re.findall('[0-9]*\.?[0-9]+元',dest_str)[0]
        numbs = re.findall('[0-9]*\.?[0-9]+',t)
        numbs = [float(i) for i in numbs ]
        dest_str = str(numbs[0])
            
    elif re.findall('[0-9]*\.?[0-9]+左右',dest_str):
        t = re.findall('[0-9]*\.?[0-9]+左右',dest_str)[0]
        numbs = re.findall('[0-9]*\.?[0-9]+',t)
        numbs = [float(i) for i in numbs ]
        dest_str = str(numbs[0])  
        
    elif re.findall('[0-9]*\.?[0-9]+上下',dest_str):
        t = re.findall('[0-9]*\.?[0-9]+上下',dest_str)[0]
        numbs = re.findall('[0-9]*\.?[0-9]+',t)
        numbs = [float(i) for i in numbs ]
        dest_str = str(numbs[0])  

    elif re.findall('价格在[0-9]*\.?[0-9]+-[0-9]*\.?[0-9]+',dest_str):
        t = re.findall('价格在[0-9]*\.?[0-9]+-[0-9]*\.?[0-9]+',dest_str)[0]
        numbs = re.findall('[0-9]*\.?[0-9]+',t)
        numbs = [float(i) for i in numbs ]
        if len(numbs)==2:
            dest_str = str((numbs[0]+numbs[1])/2.0)
        
    elif re.findall('价格在[0-9]*\.?[0-9]+',dest_str):
        t = re.findall('价格在[0-9]*\.?[0-9]+',dest_str)[0]
        numbs = re.findall('[0-9]*\.?[0-9]+',t)
        numbs = [float(i) for i in numbs ]
        dest_str = str(numbs[0])

    elif re.findall('价格[0-9]*\.?[0-9]+',dest_str):
        t = re.findall('价格[0-9]*\.?[0-9]+',dest_str)[0]
        numbs = re.findall('[0-9]*\.?[0-9]+',t)
        numbs = [float(i) for i in numbs ]
        dest_str = str(numbs[0])        
        
    elif re.findall('成交在[0-9]*\.?[0-9]+',dest_str):
        t = re.findall('成交在[0-9]*\.?[0-9]+',dest_str)[0]
        numbs = re.findall('[0-9]*\.?[0-9]+',t)
        numbs = [float(i) for i in numbs ]
        dest_str = str(numbs[0])
        
#     print(dest_str)
    return dest_str


# In[15]:


code_lis = []
code_lis_row = []
for cell in sheet1['C']:
    if cell.value and cell.value!='债券代码':
        code_lis.append(str(cell.value))
        code_lis_row.append(cell.row)
        
maxc = sheet1.max_column
    
row_lis = []
# 特殊情况
flag_hejing06 = False
remember_06row = []

flag_aoyuanzhai = False
remember_ayz_row = []
# ---------------------------------------------------------

for cell in sheet['C']:
    str_value = str(cell.value)
    # 特殊情况 金贵银业 17宜华 15宜华 18泛海
    if str_value=='002716': cell.value = '002716.SZ'
    if str_value=='101761030' or str_value=='122405': sheet['L'+str(cell.row)] = 24.95
    if str_value=='149771': sheet['L'+str(cell.row)] = 30.0
    # ---------------------------
    for i in range(len(code_lis)):
        if str_value == code_lis[i]:
            str_temp = 'L'+str(cell.row)
            sheet[str_temp].value = sheet1.cell(row =code_lis_row[i], column = maxc).value
            para = 1
            while sheet[str_temp].value is None:
                sheet[str_temp].value = sheet1.cell(row = code_lis_row[i]-para, column = maxc).value
                para += 1
            chuan = sheet[str_temp].value
            name = sheet1['B'+str(code_lis_row[i])].value
            # 特殊情况
            if name == "21番雅01":
                name = "20雅居1A"
            if name == "平裕6优":
                name = "19世茂G"
            if name == "20合景04":
                remember_04row = str(cell.row)
            if name == "20奥园01":
                remember_ay_row = str(cell.row)
            # -----------------------
            chuan = chuan.replace(';','。')
            chuan = chuan.replace('；','。')
            chuan = chuan.replace('\n','。')
            temp_list = [i.strip() for i in chuan.split('。')]
            temp_list = [i for i in temp_list if i != '']
            flag = False
            for item in temp_list:
                if name in item and ("成交价" in item or "成交在" in item):
                    sheet[str_temp].value = func1(item)
                    flag = True
                    break

            if flag == False:
                # 特殊情况
                if name == "20合景06":
                    flag_hejing06 = True
                    remember_06row.append(str(cell.row))
                if name == "21奥园债":
                    flag_aoyuanzhai = True
                    remember_ayz_row.append(str(cell.row))
                # -----------------------------------
                sheet[str_temp].value = '=f_dq_close(C'+str(sheet[str_temp].row)+',"'+str_d+'",1)'
                sheet[str_temp].fill = PatternFill(fill_type = 'solid', fgColor = 'FFFFDD')
#                 sheet[str_temp].value = '/'
            else:    
                sheet[str_temp].fill = PatternFill(fill_type = 'solid', fgColor = 'EEEEEE')
            
# 特殊情况
if flag_hejing06:
    for row in remember_06row:
        sheet['L'+row].value = sheet['L'+remember_04row].value
        
if flag_aoyuanzhai:
    for row in remember_ayz_row:
        sheet['L'+row].value = sheet['L'+remember_ay_row].value
# ----------------------------------


# In[16]:


# 调整格式
left_lis = ['A','B']
right_lis = ['C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V']

alignment = Alignment(horizontal = 'left')
for col in left_lis:
    for cell in sheet[col]:
        cell.alignment = alignment
alignment = Alignment(horizontal = 'right')
for col in right_lis:
    for cell in sheet[col]:
        cell.alignment = alignment

zimu_lis = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
width_lis = [11,15,9,13,22,14,21,19,16,14,15,9,10,10,18,14,14,14,14,7,10,10,7,7,7,7]
for i in range(len(zimu_lis)):
    sheet.column_dimensions[zimu_lis[i]].width = width_lis[i]
    
# 冻结窗格
sheet.sheet_view.topLeftCell = 'A1'
sheet.freeze_panes = 'D1'


# In[17]:


file_name1 = os.getcwd()+'/result/浮动成交收益表'+str_d[5:7]+str_d[8:10]+'.xlsx'
print("***************************************************************")
print("Output File:")
print(file_name1)
print("***************************************************************")
workbook.save(filename = file_name1)

