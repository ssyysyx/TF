#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os
import xlwings as xw
import datetime
import warnings
warnings.filterwarnings('ignore') 
from openpyxl import load_workbook

pwd = os.getcwd()

# path = ''
# for f in data_file:
#     if '产品投后管理' in f:
#         path = os.getcwd() + '/' + f
# #         print(path)

# wb = xw.Book(path)


# In[2]:


import pandas as pd
data_file = os.listdir(pwd+'/data')

print("***************************************************************")
print("Input File List:")
print(pwd+"/data/君享丰硕")
print(pwd+"/data/衍复天禄1000指增")
print(pwd+"/data/灵活对冲")
print(pwd+"/data/凡二量化")
print(pwd+"/data/赫富尊享")
print(pwd+"/data/产品投后管理")

print("***************************************************************")
print("Find and Open:")

for f in data_file:
    if "君享丰硕" in f:
        print(pwd+'/data/'+f)
        df_jxfs = pd.read_excel(pwd+'/data/'+f,sheet_name=0,header=3)
    if "衍复天禄1000指增" in f:
        print(pwd+'/data/'+f)
        df_yftl_zz = pd.read_excel(pwd+'/data/'+f,sheet_name=0)
    if "灵活对冲" in f:
        print(pwd+'/data/'+f)
        df_yf_lhdc = pd.read_excel(pwd+'/data/'+f,sheet_name=0)
    if "凡二量化" in f:
        print(pwd+'/data/'+f)
        df_felh = pd.read_excel(pwd+'/data/'+f,sheet_name=0)
    if "赫富尊享" in f:
        print(pwd+'/data/'+f)
        df_hfzx = pd.read_excel(pwd+'/data/'+f,sheet_name=0)
    if '产品投后管理' in f:
        print(pwd+'/data/'+f)
        wb = xw.Book(pwd+'/data/'+f)


# In[3]:


l = len(wb.sheets)
last_row=[int(wb.sheets[i].range('A' + str(wb.sheets[i].cells.last_cell.row)).end('up').row)  for i in range(l)]

oneday = datetime.timedelta(days=1)
LastReportDate = wb.sheets[2].range('A'+str(last_row[2])).value.date()
reportday = str(df_yftl_zz['业务日期'][0])
ReportDate = datetime.date(int(reportday[:4]),int(reportday[4:6]),int(reportday[6:]))
DateDelta = (ReportDate-LastReportDate).days

new_row = [x+DateDelta for x in last_row]


# In[4]:


# print(wb.sheets[15].name)


# In[5]:


# autofill下面一栏
for i,k in zip([2,3,4,5,6,7,8,9,10,11,14,15],['X','Z','AB','AJ','AE','AC','X','AI','X','AC','AT','F']):
    cur_rng = 'A'+str(last_row[i])+':'+ k + str(last_row[i])
    exp_rng = 'A'+str(last_row[i])+':'+ k + str(new_row[i])
    #     print(cur_rng)
    #     print(exp_rng)
    wb.sheets[i].range(cur_rng).api.AutoFill(wb.sheets[i].range(exp_rng).api,1)

# 加上日期
for d in range(1,DateDelta+1,1):
    
    for i in [2,3,4,5,6,7,8,9,10,11,14,15]:
        lastdate = wb.sheets[i].range('A'+str(last_row[i])).value.date()
        #print(lastdate)
        filldate = lastdate + datetime.timedelta(days = d)
        #print(filldate)
        wb.sheets[i].range('A'+str(last_row[i]+d)).value = filldate


# In[6]:


# 处理君享定增
jxdz_1 = df_jxfs.loc[df_jxfs['科目代码'] == '基金资产净值:','市值'].values[0]
jxdz_2 = df_jxfs.loc[df_jxfs['科目代码'] == '基金单位净值：','市值'].values[0]
jxdz_3 = df_jxfs.loc[df_jxfs['科目代码'] == '累计单位净值:','市值'].values[0]
wb.sheets[4].range('B'+str(new_row[4])).value = jxdz_2
wb.sheets[4].range('C'+str(new_row[4])).value = jxdz_3
wb.sheets[4].range('D'+str(new_row[4])).value = jxdz_1

# 处理衍复天禄1000指增
yftl_zz_1 = df_yftl_zz['客户资产净值'].iloc[0]
yftl_zz_2 = df_yftl_zz['客户资产份额'].iloc[0]
yftl_zz_3 = df_yftl_zz['虚拟计提金额'].iloc[0]
yftl_zz_4 = df_yftl_zz['单位净值'].iloc[0]
yftl_zz_5 = df_yftl_zz['累计单位净值'].iloc[0]
yftl_zz_6 = df_yftl_zz['虚拟后净值'].iloc[0]

yftl_zz_3 = yftl_zz_3.replace(',','')
yftl_zz_3 = float(yftl_zz_3)
yftl_zz_3 = "="+str(yftl_zz_3)+"-$M$367"
wb.sheets[5].range('E'+str(new_row[5])).value = yftl_zz_1
wb.sheets[5].range('F'+str(new_row[5])).value = yftl_zz_2
wb.sheets[5].range('M'+str(new_row[5])).formula = yftl_zz_3
wb.sheets[5].range('B'+str(new_row[5])).value = yftl_zz_4
wb.sheets[5].range('C'+str(new_row[5])).value = yftl_zz_5
wb.sheets[5].range('D'+str(new_row[5])).value = yftl_zz_6

# 处理衍复-灵活对冲
yf_lhdc_1 = df_yf_lhdc['单位净值'].iloc[0]
yf_lhdc_2 = df_yf_lhdc['累计单位净值'].iloc[0]
yf_lhdc_3 = df_yf_lhdc['虚拟后净值'].iloc[0]
yf_lhdc_4 = df_yf_lhdc['客户资产净值'].iloc[0]
yf_lhdc_5 = df_yf_lhdc['虚拟计提金额'].iloc[0]
wb.sheets[6].range('B'+str(new_row[6])).value = yf_lhdc_1
wb.sheets[6].range('C'+str(new_row[6])).value = yf_lhdc_2
wb.sheets[6].range('D'+str(new_row[6])).value = yf_lhdc_3
wb.sheets[6].range('E'+str(new_row[6])).value = yf_lhdc_4
wb.sheets[6].range('L'+str(new_row[6])).value = yf_lhdc_5

# 处理凡二量化对冲7号
felh_1 = df_felh['单位净值'].iloc[0]
felh_2 = df_felh['累计单位净值'].iloc[0]
felh_3 = df_felh['产品资产净值'].iloc[0]
wb.sheets[7].range('B'+str(new_row[7])).value = felh_1
wb.sheets[7].range('C'+str(new_row[7])).value = felh_2
wb.sheets[7].range('D'+str(new_row[7])).value = felh_3

# 处理赫富尊享
hfzx_1 = df_hfzx['单位净值'].iloc[0]
hfzx_2 = df_hfzx['累计单位净值'].iloc[0]
hfzx_3 = df_hfzx['虚拟后净值'].iloc[0]
hfzx_4 = df_hfzx['客户资产净值'].iloc[0]
hfzx_5 = df_hfzx['客户资产份额'].iloc[0]
hfzx_6 = df_hfzx['虚拟计提金额'].iloc[0]
wb.sheets[9].range('B'+str(new_row[9])).value = hfzx_1
wb.sheets[9].range('C'+str(new_row[9])).value = hfzx_2
wb.sheets[9].range('D'+str(new_row[9])).value = hfzx_3
wb.sheets[9].range('E'+str(new_row[9])).value = hfzx_4
wb.sheets[9].range('F'+str(new_row[9])).value = hfzx_5
wb.sheets[9].range('M'+str(new_row[9])).value = hfzx_6

# 处理产品净值图
wb.sheets[13].range('B1').value = ReportDate


# In[7]:


output_filename = pwd+"/result/产品投后管理-"+ReportDate.strftime("%Y%m%d")+".xlsx"
wb.save(output_filename)
wb.close()

# # 调整视图
# workbook = load_workbook(filename = output_filename)
# for i in [2,3,4,5,6,7,8,9,10,11,14,15]:
#     tempsheet = workbook[workbook.sheetnames[i]]
#     tempsheet.sheet_view.topLeftCell = 'A'+str(last_row[i])
    
# workbook.save(filename = output_filename)

print("***************************************************************")
print("Output File:")
print(output_filename)
print("***************************************************************")

