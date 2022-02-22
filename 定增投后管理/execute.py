#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os
import xlwings as xw
import datetime
import pandas as pd
import warnings
warnings.filterwarnings('ignore') 
from openpyxl import load_workbook

pwd = os.getcwd()
data_file = os.listdir(pwd+'/data')

print("***************************************************************")
print("Input File List:")
print(pwd+'/data/定增投后管理')
print(pwd+'/data/君享天成')
print(pwd+'/data/盯市日报')
print(pwd+'/data/收益互换日报表')

print("***************************************************************")
print("Find and Open:")
path = ''
for f in data_file:
    if '定增投后管理' in f:
        print(pwd + '/data/' + f)
        path = pwd + '/data/' + f
        df_dzth_cc = pd.read_excel(pwd+'/data/'+f,sheet_name=0)
        df_dzth_cl = pd.read_excel(pwd+'/data/'+f,sheet_name=1,header=1)
        df_dzth_jx = pd.read_excel(pwd+'/data/'+f,sheet_name=2,header=1)

workbook_dzth = load_workbook(filename = path)
wb = xw.Book(path)


# In[2]:


import pandas as pd
data_file = os.listdir(pwd+'/data')

for f in data_file:
    if "君享天成" in f:
        print(pwd + '/data/' + f)
        df_jxtc = pd.read_excel(pwd+'/data/'+f,sheet_name=0,header=3)
    if "盯市日报" in f:
        print(pwd + '/data/' + f)
        workbook_ds = load_workbook(filename = pwd+'/data/'+f)
    if "收益互换日报表" in f:
        print(pwd + '/data/' + f)
        df_hygz = pd.read_excel(pwd+'/data/'+f,sheet_name=0)
        df_bdcc = pd.read_excel(pwd+'/data/'+f,sheet_name="标的持仓")


# In[3]:


last_row_0 = wb.sheets[0].range('S'+str(wb.sheets[0].cells.last_cell.row)).end('up').row


# In[4]:


l = len(wb.sheets)
last_row=[int(wb.sheets[i].range('A' + str(wb.sheets[i].cells.last_cell.row)).end('up').row)  for i in  range(l) ]

oneday = datetime.timedelta(days=1)
LastReportDate = wb.sheets[1].range('A'+str(last_row[1])).value.date()
reportday = str(df_hygz['估值日'][0])
ReportDate = datetime.date(int(reportday[:4]),int(reportday[4:6]),int(reportday[6:8]))
DateDelta = (ReportDate-LastReportDate).days

new_row = [x+DateDelta for x in last_row]
new_row_0 = last_row_0 + DateDelta


# In[6]:


# 存公式
formula_1 = '=SUMIF(持仓!$A$2:$A$1048576,"国君互换",持仓!$F$2:$F$1048576)+SUMIF(持仓!$A$2:$A$1048576,"银河互换",持仓!$F$2:$F$1048576)'
formula_2 = '=SUMIF(持仓!$A$2:$A$1048576,"国君互换",持仓!$F$2:$F$1048576)+SUMIF(持仓!$A$2:$A$1048576,"国君互换",持仓!$N$2:$N$1048576)+SUMIF(持仓!$A$2:$A$1048576,"银河互换",持仓!$F$2:$F$1048576)+SUMIF(持仓!$A$2:$A$1048576,"银河互换",持仓!$N$2:$N$1048576)'
formula_3 = '=SUMIF(持仓!$A$2:$A$1048576,"财通产品",持仓!$F$2:$F$1048576)+SUMIF(持仓!$A$2:$A$1048576,"财通产品",持仓!$N$2:$N$1048576)'
formula_4 = '=SUM($F$2:$F$1048576)+SUM($N$2:$N$1048576)'

# 将单元格值化
temp_1 = float(df_dzth_cl['标的当前市值'].iloc[-1])
wb.sheets[1].range('D'+str(last_row[1])).value = temp_1
temp_2 = float(df_dzth_cl['标的总期初市值'].iloc[-1])
wb.sheets[1].range('C'+str(last_row[1])).value = temp_2
temp_3 = float(df_dzth_jx['资产净值'].iloc[-1])
wb.sheets[2].range('D'+str(last_row[2])).value = temp_3
temp_4 = float(df_dzth_cc['申购金额+权益收益金额'].iloc[last_row_0-2])
wb.sheets[0].range('T'+str(last_row_0)).value = temp_4


# In[7]:


# autofill下面一栏
for i,k in zip([1,2,3],['L','G','F']):
    cur_rng = 'A'+str(last_row[i])+':'+ k + str(last_row[i])
    exp_rng = 'A'+str(last_row[i])+':'+ k + str(new_row[i])
    #     print(cur_rng)
    #     print(exp_rng)
    wb.sheets[i].range(cur_rng).api.AutoFill(wb.sheets[i].range(exp_rng).api,1)

# autofill 持仓
cur_rng = 'S'+str(last_row_0)+":U"+str(last_row_0)
exp_rng = 'S'+str(last_row_0)+":U"+str(new_row_0)
wb.sheets[0].range(cur_rng).api.AutoFill(wb.sheets[0].range(exp_rng).api,1)
    
# 加上日期
for d in range(1,DateDelta+1,1):
    
    for i in [1,2,3]:
        lastdate = wb.sheets[i].range('A'+str(last_row[i])).value.date()
        #print(lastdate)
        filldate = lastdate + datetime.timedelta(days = d)
        #print(filldate)
        wb.sheets[i].range('A'+str(last_row[i]+d)).value = filldate
        
    # 第0个sheet
    lastdate = wb.sheets[0].range('S'+str(last_row_0)).value.date()
    filldate = lastdate + datetime.timedelta(days = d)
    wb.sheets[0].range('S'+str(last_row_0+d)).value = filldate
    
# 新增重置
wb.sheets[1].range('B'+str(new_row[1])).value = None


# In[8]:


# 处理持仓
wb.sheets[0].range('T'+str(new_row_0)).formula = formula_4


# In[9]:


# 处理君享天成
jxtc_1 = df_jxtc.loc[df_jxtc['科目代码'] == '基金资产净值:','市值'].values[0]
jxtc_2 = df_jxtc.loc[df_jxtc['科目代码'] == '基金单位净值：','市值'].values[0]
jxtc_3 = df_jxtc.loc[df_jxtc['科目代码'] == '累计单位净值:','市值'].values[0]

# wb.sheets[2].range('D'+str(new_row[2])).value = jxtc_1
wb.sheets[2].range('D'+str(new_row[2])).formula = formula_3
wb.sheets[2].range('B'+str(new_row[2])).value = jxtc_2
wb.sheets[2].range('C'+str(new_row[2])).value = jxtc_3


# In[10]:


# 处理盯市日报
sheet_ds = workbook_ds['Sheet1']

for cell in sheet_ds['R']:
    if cell.row>25 and (cell.row==sheet_ds.max_row or sheet_ds['R'+str(cell.row+1)].value is None):
        record_row5 = cell.row
        break
        
valueD1 = sheet_ds['R'+str(record_row5)].value
valueH1 = sheet_ds['Z'+str(record_row5)].value
valueI = sheet_ds['AB'+str(record_row5)].value


# In[11]:


# 获取盯市更新的日期commit_day 行数为record_row5-1
commitdate = sheet_ds['D'+str(record_row5-1)].value
commit_day = datetime.datetime(int(commitdate[0:4]),int(commitdate[5:7]),int(commitdate[8:10]))


# In[12]:


# 获取收益互换日报表数据 df_hygz df_bdcc
syhh_para = 11003265.09/21000006.09

syhh_data_1 = df_hygz.loc[df_hygz['交易确认书编号'] == '2020-49-01-004','未支付利率收益金额（结算货币）'].values[0]
syhh_data_2 = df_hygz.loc[df_hygz['交易确认书编号'] == '2020-49-01-003','未支付利率收益金额（结算货币）'].values[0]
syhh_data_3 = df_hygz.loc[df_hygz['交易确认书编号'] == '2020-49-01-002','未支付利率收益金额（结算货币）'].values[0]
        
valueH2 = syhh_para*syhh_data_1 + syhh_data_2 + syhh_data_3

syhh_data_4 = df_bdcc.loc[df_bdcc['证券名称'] == '光环新网','市值(计价货币)'].values[0]
syhh_data_5 = df_bdcc.loc[df_bdcc['证券名称'] == '东风股份','市值(计价货币)'].values[0]
syhh_data_6 = df_bdcc.loc[df_bdcc['证券名称'] == '东兴证券','市值(计价货币)'].values[0]

valueD2 = syhh_para*syhh_data_4 + syhh_data_5 + syhh_data_6


# In[13]:


# 填写D列和H列和I列和C列
# val = '=' + str(valueD1) + '+' + str(valueD2)
# wb.sheets[1].range('D'+str(new_row[1])).value = val
wb.sheets[1].range('D'+str(new_row[1])).formula = formula_2
wb.sheets[1].range('C'+str(new_row[1])).formula = formula_1

val = '=' + str(valueH1) + '+' + str(valueH2)
wb.sheets[1].range('H'+str(new_row[1])).value = val

wb.sheets[1].range('I'+str(new_row[1])).value = valueI


# In[14]:


# 填写补充的标的
clsy_record_row_B = None

for row in range(1,new_row[1]+1):
    if wb.sheets[1].range((row,1)).value == commit_day:
        clsy_record_row_B = row
        break
        
val = sheet_ds['K'+str(record_row5-1)].value

if clsy_record_row_B:
    wb.sheets[1].range('B'+str(clsy_record_row_B)).value = val


# In[15]:


reportplace = pwd+"/result/定增投后管理-"+ReportDate.strftime("%Y%m%d")+".xlsx"
wb.save(reportplace)
print("***************************************************************")
print("Output File:")
print(reportplace)
print("***************************************************************")
wb.close()


# In[ ]:




